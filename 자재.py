import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from io import BytesIO

# 설정
FILE_NAME = "자재승인현황.xlsx"
PDF_DIR = "pdf_storage"
START_ROW = 5

if not os.path.exists(PDF_DIR):
    os.makedirs(PDF_DIR)

st.set_page_config(page_title="자재 관리 시스템", layout="wide")

# 사이드바 메뉴
menu = st.sidebar.radio("📌 메뉴 선택", ["데이터 입력 (항목1)", "결재본 조회 및 다운로드 (항목2)"])

# ---------------------------------------------------------
# [항목1] 데이터 입력 섹션
# ---------------------------------------------------------
if menu == "데이터 입력 (항목1)":
    st.title("🏗️ 자재 승인 정보 및 PDF 등록")
    
    file_exists = os.path.exists(FILE_NAME)
    
    if not file_exists:
        st.warning("⚠️ 아직 등록된 엑셀 양식이 없습니다. 아래에서 파일을 먼저 한 번만 업로드해 주세요.")
        uploaded_excel = st.file_uploader("엑셀 양식 파일 업로드", type="xlsx")
        if uploaded_excel:
            with open(FILE_NAME, "wb") as f:
                f.write(uploaded_excel.getbuffer())
            st.success("✅ 엑셀 양식이 등록되었습니다!")
            st.rerun()
    else:
        st.info(f"📂 현재 '{FILE_NAME}' 파일을 사용 중입니다.")
        if st.button("다른 엑셀 파일로 교체하기"):
            os.remove(FILE_NAME)
            st.rerun()

    if file_exists:
        pdf_file = st.file_uploader("해당 순번 결재본 PDF 업로드", type="pdf")

        with st.form("input_form"):
            target_no = st.number_input("입력할 순번 (A열 번호)", min_value=1, max_value=100, value=1)
            
            col1, col2 = st.columns(2)
            with col1:
                req_date = st.date_input("요청일자")
                item_name = st.text_input("품명")
            with col2:
                app_date = st.date_input("승인일자")
                company = st.text_input("제조회사명")
                
            c1, c2, c3 = st.columns(3)
            with c1: is_ks = st.checkbox("KS")
            with c2: is_env = st.checkbox("환경표지")
            with c3: is_gr = st.checkbox("GR")

            submitted = st.form_submit_button("엑셀 저장 및 PDF 등록")

        if submitted:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            actual_row = target_no + (START_ROW - 1)
            
            f_req_date = req_date.strftime("'%y.%m.%d")
            f_app_date = app_date.strftime("'%y.%m.%d")
            
            data = [f_req_date, f_app_date, item_name, company, 
                    '○' if is_ks else '', '○' if is_env else '', '○' if is_gr else '']
            
            for i, val in enumerate(data, start=2):
                ws.cell(row=actual_row, column=i, value=val)
            
            wb.save(FILE_NAME)
            st.success(f"✅ {target_no}번 데이터가 엑셀에 저장되었습니다!")

            if pdf_file:
                pdf_path = os.path.join(PDF_DIR, f"{target_no}.pdf")
                with open(pdf_path, "wb") as f:
                    f.write(pdf_file.getbuffer())
                st.info(f"📎 {target_no}.pdf 결재본이 등록되었습니다.")

# ---------------------------------------------------------
# [항목2] 결재본 조회 및 다운로드 섹션 (엑셀 다운로드 추가됨!)
# ---------------------------------------------------------
elif menu == "결재본 조회 및 다운로드 (항목2)":
    st.title("📂 파일 통합 조회 및 다운로드")
    
    # --- 1. 전체 엑셀 파일 다운로드 ---
    st.subheader("📊 전체 엑셀 현황 내려받기")
    if os.path.exists(FILE_NAME):
        with open(FILE_NAME, "rb") as f:
            st.download_button(
                label="📥 수정된 전체 엑셀 파일 다운로드",
                data=f,
                file_name="최종_자재승인현황.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("아직 저장된 엑셀 파일이 없습니다.")

    st.divider()

    # --- 2. 개별 PDF 조회 ---
    st.subheader("📄 순번별 결재본 PDF 찾기")
    search_no = st.number_input("조회할 순번 입력", min_value=1, max_value=100, value=1)
    file_path = os.path.join(PDF_DIR, f"{search_no}.pdf")
    
    if os.path.exists(file_path):
        st.success(f"📄 {search_no}번 결재본을 찾았습니다.")
        with open(file_path, "rb") as f:
            st.download_button(
                label=f"📥 {search_no}번 PDF 다운로드",
                data=f,
                file_name=f"결재본_순번_{search_no}.pdf",
                mime="application/pdf"
            )
    else:
        st.warning(f"⚠️ {search_no}번으로 등록된 PDF 파일이 없습니다.")